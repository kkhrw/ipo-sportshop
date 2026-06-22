from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator
from django.core.mail import EmailMessage
from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User
from django import forms
from django.core.paginator import Paginator

from rest_framework import viewsets, permissions, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser

from .models import Product, Category, Manufacturer, Cart, CartItem, Order, OrderItem, Profile
from .forms import CheckoutForm
from .serializers import (
    CategorySerializer, ManufacturerSerializer, ProductSerializer,
    CartSerializer, CartItemSerializer, OrderSerializer, OrderItemSerializer,
    UserSerializer, ProfileSerializer
)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io


class RegisterForm(UserCreationForm):
    email = forms.EmailField(required=True)
    full_name = forms.CharField(max_length=200, required=False)
    phone = forms.CharField(max_length=20, required=False)

    class Meta:
        model = User
        fields = ('username', 'email', 'password1', 'password2')

    def save(self, commit=True):
        user = super().save(commit)
        if commit and hasattr(user, 'profile'):
            user.profile.full_name = self.cleaned_data.get('full_name', '')
            user.profile.phone = self.cleaned_data.get('phone', '')
            user.profile.save()
        return user


def home_page(request):
    popular_products = Product.objects.all()[:6]
    categories = Category.objects.all()
    context = {
        'popular_products': popular_products,
        'categories': categories,
    }
    return render(request, 'shop/index.html', context)


def about_page(request):
    return render(request, 'shop/about.html')


def author_page(request):
    return render(request, 'shop/author.html')


def product_list(request):
    products = Product.objects.all()
    
    category_id = request.GET.get('category')
    if category_id and category_id != '':
        products = products.filter(category_id=category_id)
    else:
        category_id = None
    
    manufacturer_id = request.GET.get('manufacturer')
    if manufacturer_id and manufacturer_id != '':
        products = products.filter(manufacturer_id=manufacturer_id)
    else:
        manufacturer_id = None
    
    search_query = request.GET.get('search')
    if search_query and search_query != 'None' and search_query.strip():
        products = products.filter(
            Q(name__icontains=search_query) | 
            Q(description__icontains=search_query)
        )
    else:
        search_query = None
    
    paginator = Paginator(products, 9)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    categories = Category.objects.all()
    manufacturers = Manufacturer.objects.all()
    
    context = {
        'page_obj': page_obj,
        'categories': categories,
        'manufacturers': manufacturers,
        'current_category': category_id,
        'current_manufacturer': manufacturer_id,
        'search_query': search_query,
    }
    
    return render(request, 'shop/catalog.html', context)


def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    context = {'product': product}
    return render(request, 'shop/product_detail.html', context)


@login_required
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_item, created = CartItem.objects.get_or_create(
        cart=cart, product=product, defaults={'quantity': 1}
    )
    if not created:
        if cart_item.quantity < product.stock_quantity:
            cart_item.quantity += 1
            cart_item.save()
            messages.success(request, f'Количество товара "{product.name}" увеличено в корзине')
        else:
            messages.warning(request, f'Нельзя добавить больше товара "{product.name}". На складе: {product.stock_quantity} шт.')
    else:
        messages.success(request, f'Товар "{product.name}" добавлен в корзину')
    return redirect('shop:cart_view')


@login_required
def update_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, pk=item_id, cart__user=request.user)
    if request.method == 'POST':
        quantity = int(request.POST.get('quantity', 1))
        if quantity > cart_item.product.stock_quantity:
            messages.error(request, f'Количество не может превышать {cart_item.product.stock_quantity} (доступно на складе)')
        elif quantity < 1:
            messages.error(request, 'Количество должно быть не меньше 1')
        else:
            cart_item.quantity = quantity
            cart_item.save()
            messages.success(request, f'Количество товара "{cart_item.product.name}" обновлено')
    return redirect('shop:cart_view')


@login_required
def remove_from_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, pk=item_id, cart__user=request.user)
    product_name = cart_item.product.name
    cart_item.delete()
    messages.success(request, f'Товар "{product_name}" удален из корзины')
    return redirect('shop:cart_view')


@login_required
def cart_view(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_items = cart.items.all()
    total_cost = cart.total_cost()
    context = {
        'cart': cart,
        'cart_items': cart_items,
        'total_cost': total_cost,
    }
    return render(request, 'shop/cart.html', context)


def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, f'Добро пожаловать, {user.username}!')
            next_url = request.GET.get('next', 'shop:home')
            return redirect(next_url)
        else:
            messages.error(request, 'Неверное имя пользователя или пароль')
    return render(request, 'shop/login.html')


def user_register(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            try:
                user = form.save()
                login(request, user)
                messages.success(request, f'Регистрация успешна! Добро пожаловать, {user.username}!')
                return redirect('shop:home')
            except Exception as e:
                messages.error(request, f'Ошибка: {str(e)}')
        else:
            # Показываем все ошибки формы
            errors = form.errors.as_text()
            messages.error(request, f'Ошибка в форме: {errors}')
            print("Ошибки формы:", form.errors)
    else:
        form = RegisterForm()
    return render(request, 'shop/register.html', {'form': form})

@login_required
def user_logout(request):
    logout(request)
    messages.success(request, 'Вы успешно вышли из системы')
    return redirect('shop:home')


@login_required
def profile_view(request):
    return render(request, 'shop/profile.html')


@login_required
def checkout(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_items = cart.items.all()
    
    if not cart_items.exists():
        messages.warning(request, 'Ваша корзина пуста! Добавьте товары перед оформлением заказа.')
        return redirect('shop:product_list')
    
    total_cost = cart.total_cost()
    
    if request.method == 'POST':
        form = CheckoutForm(request.POST)
        if form.is_valid():
            delivery_address = form.cleaned_data['delivery_address']
            customer_email = form.cleaned_data['email']
            
            order = Order.objects.create(
                user=request.user,
                delivery_address=delivery_address,
                total_cost=total_cost
            )
            
            for cart_item in cart_items:
                OrderItem.objects.create(
                    order=order,
                    product=cart_item.product,
                    quantity=cart_item.quantity,
                    price=cart_item.product.price
                )
            
            excel_file = generate_receipt_excel(order)
            email_sent = send_receipt_email(order, customer_email, excel_file)
            
            if email_sent:
                order.email_sent = True
                order.save()
                messages.success(request, f'Заказ #{order.id} успешно оформлен! Чек отправлен на {customer_email}')
            else:
                messages.warning(request, f'Заказ #{order.id} оформлен, но не удалось отправить чек на email.')
            
            cart_items.delete()
            return redirect('shop:home')
    else:
        form = CheckoutForm(initial={'email': request.user.email})
    
    context = {
        'form': form,
        'cart_items': cart_items,
        'total_cost': total_cost,
    }
    return render(request, 'shop/checkout.html', context)


def generate_receipt_excel(order):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Чек заказа"
    
    header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='667EEA', end_color='764BA2', fill_type='solid')
    title_font = Font(name='Arial', size=16, bold=True, color='667EEA')
    normal_font = Font(name='Arial', size=11)
    bold_font = Font(name='Arial', size=11, bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = 'Zona Sporta - Чек заказа'
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center')
    
    ws['A3'] = 'Номер заказа:'
    ws['B3'] = f'#{order.id}'
    ws['A4'] = 'Дата:'
    ws['B4'] = order.created_at.strftime('%d.%m.%Y %H:%M')
    ws['A5'] = 'Покупатель:'
    ws['B5'] = order.user.username
    ws['A6'] = 'Адрес доставки:'
    ws['B6'] = order.delivery_address
    
    for row in range(3, 7):
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'].font = normal_font
    
    ws['A8'] = ''
    
    headers = ['№', 'Наименование товара', 'Кол-во', 'Цена (руб.)', 'Сумма (руб.)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row_num = 10
    for index, item in enumerate(order.items.all(), 1):
        ws.cell(row=row_num, column=1, value=index).border = thin_border
        ws.cell(row=row_num, column=2, value=item.product.name).border = thin_border
        ws.cell(row=row_num, column=3, value=item.quantity).border = thin_border
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=4, value=float(item.price)).border = thin_border
        ws.cell(row=row_num, column=4).number_format = '#,##0.00'
        ws.cell(row=row_num, column=5, value=float(item.item_cost())).border = thin_border
        ws.cell(row=row_num, column=5).number_format = '#,##0.00'
        row_num += 1
    
    ws.cell(row=row_num + 1, column=4, value='ИТОГО:').font = bold_font
    ws.cell(row=row_num + 1, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=row_num + 1, column=5, value=float(order.total_cost)).font = bold_font
    ws.cell(row=row_num + 1, column=5).number_format = '#,##0.00'
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def send_receipt_email(order, email, excel_file):
    try:
        subject = f'Чек заказа #{order.id} - Zona Sporta'
        message = f"""Здравствуйте, {order.user.username}!

Спасибо за ваш заказ в нашем магазине спортивных товаров!

Детали заказа:
• Номер заказа: #{order.id}
• Дата: {order.created_at.strftime('%d.%m.%Y %H:%M')}
• Адрес доставки: {order.delivery_address}
• Общая сумма: {order.total_cost} руб.

К чеку прикреплён файл в формате Excel с подробной информацией о заказе.

С уважением,
Команда магазина спортивных товаров
"""
        
        email_message = EmailMessage(
            subject=subject,
            body=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[email]
        )
        
        filename = f'чек_заказа_{order.id}_{order.created_at.strftime("%Y%m%d")}.xlsx'
        email_message.attach(filename, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        email_message.send()
        return True
    except Exception as e:
        print(f'Ошибка отправки email: {e}')
        return False


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'description']
    ordering_fields = ['name']


class ManufacturerViewSet(viewsets.ModelViewSet):
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'country']
    ordering_fields = ['name', 'country']


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'description', 'category__name', 'manufacturer__name']
    ordering_fields = ['name', 'price', 'created_at']

    
    @action(detail=False, methods=['get'])
    def by_category(self, request):
        category_id = request.query_params.get('category_id')
        if category_id:
            products = self.queryset.filter(category_id=category_id)
            serializer = self.get_serializer(products, many=True)
            return Response(serializer.data)
        return Response({'error': 'Укажите category_id'}, status=400)

    @action(detail=False, methods=['get'])
    def by_manufacturer(self, request):
        manufacturer_id = request.query_params.get('manufacturer_id')
        if manufacturer_id:
            products = self.queryset.filter(manufacturer_id=manufacturer_id)
            serializer = self.get_serializer(products, many=True)
            return Response(serializer.data)
        return Response({'error': 'Укажите manufacturer_id'}, status=400)
    

class CartViewSet(viewsets.ModelViewSet):
    serializer_class = CartSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Cart.objects.filter(user=self.request.user)

    @action(detail=False, methods=['get'])
    def my_cart(self, request):
        cart, created = Cart.objects.get_or_create(user=request.user)
        serializer = self.get_serializer(cart)
        return Response(serializer.data)


class CartItemViewSet(viewsets.ModelViewSet):
    serializer_class = CartItemSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return CartItem.objects.filter(cart__user=self.request.user)

    def perform_create(self, serializer):
        cart, created = Cart.objects.get_or_create(user=self.request.user)
        serializer.save(cart=cart)


class OrderViewSet(viewsets.ModelViewSet):
    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return OrderSerializer
        return OrderSerializer

    def get_queryset(self):
        user = self.request.user
        if user.is_staff:
            return Order.objects.all()
        return Order.objects.filter(user=user)


class ProfileViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def list(self, request):
        profile = request.user.profile
        serializer = ProfileSerializer(profile)
        return Response(serializer.data)

    def partial_update(self, request):
        profile = request.user.profile
        serializer = ProfileSerializer(profile, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAdminUser]

    @action(detail=False, methods=['get'])
    def me(self, request):
        serializer = UserSerializer(request.user)
        return Response(serializer.data)

@login_required
def profile_view(request):
    return render(request, 'shop/profile.html')